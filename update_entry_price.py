import pandas as pd
from pytdx.hq import TdxHq_API
from pytdx.config.hosts import hq_hosts
from datetime import datetime
import random

# Configuration
STOCK_FILE = "output/stock_tracking_20260414.csv"
TRADE_DATE = 20260414
TARGET_TIME = "13:05:00"  # 通达信5分钟K线时间戳为K线结束时间，13:05的K线open即下午13:00开盘价

# Read original data
df = pd.read_csv(STOCK_FILE)
original_df = df.copy()
print("Original tracking data:")
print(original_df[["代码", "名称", "参考入场价(2026-04-14)", "累计涨跌幅(%)"]])
print("\n" + "="*80 + "\n")

# Initialize TDX API
api = TdxHq_API()
success_count = 0
fail_count = 0

# Try random servers from the built-in list
random.shuffle(hq_hosts)
connected = False
for host in hq_hosts[:10]:
    try:
        print(f"Trying server {host[1]}:{host[2]}...")
        if api.connect(host[1], host[2], time_out=3):
            connected = True
            break
    except Exception as e:
        print(f"Failed to connect to {host[1]}:{host[2]}: {str(e)}")
        continue

if not connected:
    print("All servers failed to connect")
    exit(1)

print(f"Connected to TDX server {host[1]}:{host[2]} successfully\n")
    
for idx, row in df.iterrows():
        code_full = row["代码"]
        code, market_suffix = code_full.split(".")
        market = 1 if market_suffix == "SH" else 0
        
        print(f"Processing {code_full} {row['名称']}...")
        
        # Get 5min kline data for 2026-04-14
        klines_5min = api.get_security_bars(
            category=0,  # 5min kline
            market=market,
            code=code,
            start=0,
            count=48  # full day 48 bars of 5min
        )
        
        entry_price = None
        # Find 13:05 bar (通达信5分钟K线为结束时间，此bar的open即13:00开盘价)
        if klines_5min:
            for k in klines_5min:
                # Format datetime from pytdx fields
                kline_date = f"{k['year']:04d}{k['month']:02d}{k['day']:02d}"
                kline_time_str = f"{k['hour']:02d}:{k['minute']:02d}:00"
                if kline_date == str(TRADE_DATE) and kline_time_str == TARGET_TIME:
                    entry_price = k["open"]
                    print(f"  Found 13:00 opening price (from 13:05 K线): {entry_price:.2f}")
                    break
        
        # Fallback to daily close price if 13:05 K线价格未找到
        if entry_price is None:
            print(f"  13:05 K线价格未找到，获取日收盘价作为替代...")
            daily_kline = api.get_security_bars(
                category=9,  # daily kline
                market=market,
                code=code,
                start=0,
                count=10
            )
            for k in daily_kline:
                kline_date = f"{k['year']:04d}{k['month']:02d}{k['day']:02d}"
                if kline_date == str(TRADE_DATE):
                    entry_price = k["close"]
                    print(f"  Using daily close price: {entry_price:.2f}")
                    break
        
        if entry_price is not None:
            # Update entry price
            df.at[idx, "参考入场价(2026-04-14)"] = round(entry_price, 2)
            # Calculate cumulative return if latest price exists
            if pd.notna(row["2026-04-15最新价"]) and row["2026-04-15最新价"] > 0:
                cum_return = (row["2026-04-15最新价"] / entry_price - 1) * 100
                df.at[idx, "累计涨跌幅(%)"] = round(cum_return, 2)
            success_count += 1
        else:
            print(f"  Failed to get price for {code_full}")
            fail_count += 1
    
api.disconnect()

print("\n" + "="*80 + "\n")
print(f"Processing completed: {success_count} success, {fail_count} fail\n")

# Print comparison
comparison = pd.DataFrame({
    "代码": df["代码"],
    "名称": df["名称"],
    "原入场价": original_df["参考入场价(2026-04-14)"],
    "新入场价": df["参考入场价(2026-04-14)"],
    "原累计涨跌幅(%)": original_df["累计涨跌幅(%)"],
    "新累计涨跌幅(%)": df["累计涨跌幅(%)"]
})
print("Update comparison:")
print(comparison)

# Save updated data
output_file = "output/stock_tracking_20260414_updated.csv"
df.to_csv(output_file, index=False, encoding="utf-8-sig")
print(f"\nUpdated tracking table saved to {output_file}")

# Also update the Excel file
from openpyxl import load_workbook
excel_file = "output/stock_pool_tracking_total.xlsx"
wb = load_workbook(excel_file)
if "20260414" in wb.sheetnames:
    ws = wb["20260414"]
    # Update column C (reference entry price) and column G (cumulative return)
    for idx, row in df.iterrows():
        ws.cell(row=idx+2, column=3, value=row["参考入场价(2026-04-14)"])
        ws.cell(row=idx+2, column=7, value=row["累计涨跌幅(%)"])
    wb.save(excel_file)
    print(f"Excel tracking sheet {excel_file} (sheet 20260414) updated successfully")
